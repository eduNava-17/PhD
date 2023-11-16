    # -*- coding: utf-8 -*-
"""
Created on Thu Oct 15 14:37:43 2020

@author: J Edu
"""

""" DESIREE Analysis """

from sys import path
import os.path
'append the path where you have the const, func nptdms scripts'
# path.append('/Users/Edu/Documents/Python scripts/Analysis program/')
import scipy.constants as const
import func as func
# import converttdms
import matplotlib as mpl
from nptdms import TdmsFile, TdmsGroup
import numpy as np
from time import time
from statsmodels.stats.weightstats import DescrStatsW
import logging
import csv
import h5py
import pickle
import func as fn
import matplotlib.pyplot as plt
from mpl_toolkits.axes_grid1 import make_axes_locatable
from numba import jit
from numba import vectorize, float64
import threading
import queue
from periodictable import formula
import datetime
import openpyxl


'''
This script works with the nptdms version 1.1.0
'''

mpl.rcParams['agg.path.chunksize'] = 10000


class Analysis:
    LE_LEBC = 3.25 #m
    HE_HEBC = 6.71
    # the first for S and the second for A ring
    LEBC_ring = [7.24,9.59] #m
    HEBC_ring = [9.93,7.80]
    ring_circ = [8.68,8.71] # m
    rad = ring_circ[0]/(2*np.pi) #m
    cen_RAES = 0
    light_speed = 299792458
    dia = str(datetime.date.today())

    def __init__(self, bins, source_path, saving_path, _channels, method, ini_file, total_counts):
        self.DES_path = source_path
        self.save_path = saving_path
        self.ini_file = ini_file
        self.method = method
        self.laser_on = False
        self.bins = bins
        self.counter_LPTS = 0
        # self.total_counts_per_wavelength = np.zeros(wl_length) #super specific for this case!!
        self.total_counts = total_counts
        if not os.path.exists(self.save_path):
            os.makedirs(self.save_path)
            print("Directory " , self.save_path,  " Created ")
        else:    
            print("Directory " , self.save_path ,  " already exists")
        self.channel_names = []
        self.channels = []
        self.No_channels = len(_channels)
        self.tdms_0()
        # for ch in _channels:
            # self.channels.append(self.channel_names.index(ch))
        print ('You are in the right script...')

    def ini_params(self,exp, _file,start,stop,hdf5_groupname, n1,n2,skips,counter,wls,b_fac,b_range,countsON = False, OPO_Pow = False):
        self.file_name = _file
        self.experiment = exp
        self.start = start
        self.stop = stop
        self.cycle_max = n2
        self.cycle_min = n1
        self.iscountsOn = countsON
        self.isOPOon = OPO_Pow
        self.counter = counter
        self.l = counter
        self.skips = skips
        self.savePath = self.save_path + self.experiment + self.file_name + '.hdf'
        self.tdms_path = self.DES_path + self.experiment + self.file_name + '.tdms'
        
        if self.method == 'linear':
            self.Nbins = int(abs(start - stop)/self.bins)
        elif self.method == 'log':
            self.Nbins = self.bins
        if hdf5_groupname:
            print ('In hdf5 mode')
            self.hdf5_groupname = 1
            if not os.path.exists(self.savePath):
                self.hdf5()
                print ('I have created the hdf5, you can proceed with .txt conversion')
                self.hdf5_LPTS_3()
                # answer = input('no wavelength 0, one 1, or multiple wavelengths 2, Gustav multiwavelength method 3, LPTS method 4, LPTS method 5  ? ')
                # if answer == '0':
                #     self.hdf5_handler()
                # elif answer == '1':
                #     self.hdf5_handler()
                # elif answer == '2':
                #     self.hdf5_wavelengths(wls)
                # elif answer == '3':
                #     self.hdf5_wavelengths_Gustav(wls,skips)
                # elif answer == '4':
                #     self.hdf5_wavelengths_LPTS_2(wls,skips,step_wv)
                # elif answer == '5':
                #     self.hdf5_LPTS_2(wls,b_fac,b_range)
            elif os.path.exists(self.savePath):
                self.is_wavelengths_ready = False
                if self.is_wavelengths_ready:
                    self.hdf5_LPTS_3()
                    wls = self.wavelengths
                print ('hdf5 already exists :) , proceeding with .txt conversion')
                print ('Are you okay there? ')
                # answer = input('no wavelength 0, one 1, or multiple wavelengths 2, Gustav multiwavelength method 3, LPTS method 4, LPTS method 5 , old LPTS method 6? ')
                answer = '6'
                if answer == '0':
                    self.hdf5_handler()
                elif answer == '1':
                    self.hdf5_handler()
                elif answer == '2':
                    self.hdf5_wavelengths(wls)
                elif answer == '3':
                    self.hdf5_wavelengths_Gustav(wls,skips)
                elif answer == '4':
                    self.hdf5_LPTS_3()
                elif answer == '5':
                    self.hdf5_LPTS_2(wls,b_fac,b_range)
                elif answer == '6':
                    self.hdf5_pumpingStates(wls,b_fac,b_range)
                # elif answer == '6':
                #     self.hdf5_LPTS(wls,step_wv)
        
        
    def ini_params2(self,_file,start,stop,hdf5_groupname, n1,n2,skips,counter,wls,countsON = False, OPO_Pow = False):
        self.file_name = _file
        self.start = start
        self.stop = stop
        self.cycle_max = n2
        self.cycle_min = n1
        self.iscountsOn = countsON
        self.isOPOon = OPO_Pow
        self.counter = counter
        self.l = counter
        self.skips = skips
        self.savePath = self.save_path + self.file_name + '.hdf'
        self.tdms_path = self.DES_path + self.file_name + '.tdms'
        
        if self.method == 'linear':
            self.Nbins = int(abs(start - stop)/self.bins)
        elif self.method == 'log':
            self.Nbins = self.bins
        if hdf5_groupname:
            print ('In hdf5 mode')
            self.hdf5_groupname = 1
            if not os.path.exists(self.savePath):
                self.hdf5()
                print ('I have created the hdf5, you can proceed with .txt conversion')
                self.hdf5_LPTS_3()
                # answer = input('no wavelength 0, one 1, or multiple wavelengths 2, Gustav multiwavelength method 3, LPTS method 4, LPTS method 5  ? ')
                # if answer == '0':
                #     self.hdf5_handler()
                # elif answer == '1':
                #     self.hdf5_handler()
                # elif answer == '2':
                #     self.hdf5_wavelengths(wls)
                # elif answer == '3':
                #     self.hdf5_wavelengths_Gustav(wls,skips)
                # elif answer == '4':
                #     self.hdf5_wavelengths_LPTS_2(wls,skips,step_wv)
                # elif answer == '5':
                #     self.hdf5_LPTS_2(wls,b_fac,b_range)
            elif os.path.exists(self.savePath):
                self.is_wavelengths_ready = False
                if self.is_wavelengths_ready:
                    self.hdf5_LPTS_3()
                    wls = self.wavelengths
                print ('hdf5 already exists :) , proceeding with .txt conversion')
                print ('Are you okay there? ')
                # answer = input('no wavelength 0, one 1, or multiple wavelengths 2, Gustav multiwavelength method 3, LPTS method 4, LPTS method 5 , old LPTS method 6? ')
                answer = '6'
                if answer == '0':
                    self.hdf5_handler()
                elif answer == '1':
                    self.hdf5_handler()
                elif answer == '2':
                    self.hdf5_wavelengths(wls)
                elif answer == '3':
                    self.hdf5_wavelengths_Gustav(wls,skips)
                elif answer == '4':
                    self.hdf5_LPTS_3()
                elif answer == '5':
                    self.hdf5_LPTS_2(wls)
                elif answer == '6':
                    self.hdf5_pumpingStates(wls)
                # elif answer == '6':
                #     self.hdf5_LPTS(wls,step_wv)
        elif not hdf5_groupname:
            print ('In TDMS mode', _file)
            # self.tdms_handler(skips)
            self.hdf5_pumpingStates(wls)
            # self.Signal_vs_power()
            # if self.l != 50:
                # print ('Entering tdms_handler...', _file)
                # self.tdms_handler(skips)
            # elif self.l ==50:
                # print ('inside tdms_handler_photons_current, \n')
                # self.tdms_handler_photons_current(skips)
                
    def ini_params_singleShot(self,exp, _file,start,stop,hdf5_groupname, n1,n2,cycles,skips,counter,wls,step_wv,ini_final_pows,countsON = False, OPO_Pow = True):
        self.file_name = _file
        self.experiment = exp
        self.start = start
        self.stop = stop
        self.l = counter
        self.cycle_max = n2
        self.cycle_min = n1
        self.iscountsOn = countsON
        self.isOPOon = OPO_Pow
        self.savePath = self.save_path + self.experiment + self.file_name + '.hdf'
        self.tdms_path = self.DES_path + self.experiment + self.file_name + '.tdms'
        if self.method == 'linear':
            self.Nbins = int(abs(start - stop)/self.bins)
        elif self.method == 'log':
            self.Nbins = self.bins
        if hdf5_groupname:
            print ('In hdf5 mode')
            self.hdf5_groupname = 1
            if not os.path.exists(self.savePath):
                self.hdf5()
                print ('I have created the hdf5, now proceeding with .txt conversion')
                answer = input('no wavelength 0, one 1, or multiple wavelengths 2, Gustav multiwavelength method 3, LPTS method 4  ? ')
                if answer == '0':
                    self.hdf5_handler()
                elif answer == '1':
                    self.hdf5_handler()
                elif answer == '2':
                    self.hdf5_wavelengths(wls)
                elif answer == '3':
                    self.hdf5_wavelengths_Gustav(wls,skips)
                elif answer == '4':
                    self.hdf5_wavelengths_LPTS_2(wls,cycles,skips,step_wv,ini_final_pows)
            elif os.path.exists(self.savePath):
                print ('hdf5 already exists :) , proceeding with .txt conversion')
                answer = input('no wavelength 0, one 1, or multiple wavelengths 2, Gustav multiwavelength method 3, LPTS method 4 ? ')
                if answer == '0':
                    self.hdf5_handler()
                elif answer == '1':
                    self.hdf5_handler()
                elif answer == '2':
                    self.hdf5_wavelengths(wls)
                elif answer == '3':
                    self.hdf5_wavelengths_Gustav(wls,skips)
                elif answer == '4':
                    self.hdf5_wavelengths_LPTS_2(wls,cycles,skips,step_wv,ini_final_pows)

        elif not hdf5_groupname:
            print ('In TDMS mode', _file)
            if self.l != 50:
                print ('Entering tdms_handler...', _file)
                self.tdms_handler(cycles,skips)
            elif self.l ==50:
                print ('inside tdms_handler_photons_current, \n')
                self.tdms_handler_photons_current(cycles,skips)
        
    ### this will set the channel numbers ########
    def tdms_0(self):
        print ('list of channels: ' )
        with TdmsFile.open(self.DES_path + self.ini_file + '.tdms') as tdms_file:
            self.group = tdms_file.groups()
            print ('This are the properties: ', tdms_file.properties)
            for gr in self.group:
                if len(gr.properties) != 0:
            # print ('properties for the group: ', gr.properties.values())
                    channel_x = gr.properties['Description']
            # print (channel_x)
                    self.channel_names.append(channel_x)

    def hdf5(self):
        btime = time()
        with TdmsFile.open(self.tdms_path) as tdms_file:
            tdms_file.as_hdf(self.savePath, 'w', group = str(self.hdf5_groupname))
        print ('hdf5 creation' + ' took : ' + str(round(time() - btime,4)) + ' s ')
        
    def energy_search(self,_file,wavelengths):
        print ('length of wavelength list: ', len(wavelengths))
        with TdmsFile.open(self.tdms_path) as tdms_file:
            group = tdms_file.groups()
            ch = self.channels[0]
            ens = group[ch]['Pulse Energy (uJ)']
            energy = ens[:]
            wl_energies = []
            for w in wavelengths:
                wls_ = group[ch]['Wavelength (nm)']
                wls = wls_[:]
                wls = np.array(wls)
                w_list = np.where(wls == w)[0]
                sum_energies = []
                for j,val in enumerate(w_list):
                    sum_energies.append(energy[val])
                    if j > len(w_list)-1 :
                        break
                wl_energies.append([w, np.average(sum_energies)])
                print ('current wl: ', w)
            self.save_energies(wl_energies, 'wavelengths_analysis/')
            # print(dir(groups[ch]))
            # path = groups[ch].path()

                
    def energy_wl(self,_file,wavelength):
        with TdmsFile.open(self.tdms_path) as tdms_file:
            group = tdms_file.groups()
            ch = self.channels[0]
            ens = group[ch]['Pulse Energy (uJ)']
            energy = ens[:]
            wls_ = group[ch]['Wavelength (nm)']
            wls = wls_[:]
            wls = np.array(wls)
            w_list = np.where(wls == wavelength)[0]
            sum_energies = []
            for j,val in enumerate(w_list):
                sum_energies.append(energy[val])
                if j > len(w_list)-1 :
                    break
            wl_energy = np.average(sum_energies)
            return wl_energy
            
    def tdms_handler(self,skips):
        self.atime = time()
        if os.path.exists(self.tdms_path):
            with TdmsFile.open(self.tdms_path) as tdms_file:
                print (tdms_file.properties)
                group= tdms_file.groups()
                yields = []
                parsback_values = []
                count_rate = []
                for l,i in enumerate(self.channels):
            #     # if gr.properties['Description'] == self.channel_names[self.No_Chan[0]]: # the name of the channel might change
                    channel = group[i]['Time (s)']
                    n_cycle = group[i]['Cycle No.']
                    n_c = len(channel)
                    m_c = round(max(channel),6)
                    print ('from file: ', self.file_name[-7:], ', length of channel: ', i ,n_c, 'cycle time (rounded):', m_c)
                    self.cycle_data = np.array(n_cycle[:])
                    self.No_cycles = max(n_cycle)
                    # elem1 = self.cycle_data.tolist().index(cycles[1])
                    # elem0 = self.cycle_data.tolist().index(cycles[0])
                    if self.cycle_max == 'max':                        
                        self.data = channel[:]
                        print ('No cycles is:', self.No_cycles, 'length of time list:', len(self.data))
                    # elif self.cycle_max == 'slice':
                            # self.data = channel[elem0:elem1]
                    else:
                        top_cycle = self.cycle_data.tolist().index(self.cycle_max)
                        self.data = channel[:top_cycle]
                        print ('top_cycle index is:', top_cycle, 'length of time list:', len(self.data))

                    yields.append(self.bin_data())
                    parsback_values.append([self.file_name, self.No_cycles, n_c,m_c,0,self.x_mid,self.l,self.start,self.stop]) # file_name, length of channel, #No.cycles,storage_time,wl,firstpointtim
                if self.isOPOon:
                    self.save_data_OPO_pow(yields,parsback_values)
                else:
                    self.save_data(yields,parsback_values,0,0)
                # if self.iscountsOn:
                #     count_rate.append(self.countRate())
                #     self.save_data(yields, parsback_values,count_rate,0)
                # elif not self.iscountsOn:
                #     count_rate = 0
                #     self.save_data(yields, parsback_values,count_rate,0)
        else:
            print ('file: ', self.tdms_path, ' does not exist in that folder')
                    
        print (' Handling took : ' + str(round(time() - self.atime,4)) + ' s ')

    def Signal_vs_power(self):
        self.atime = time()
        with TdmsFile.open(self.tdms_path) as tdms_file:
            # print (tdms_file.properties)
            group= tdms_file.groups()
        #     # if gr.properties['Description'] == self.channel_names[self.No_Chan[0]]: # the name of the channel might change
            times = group[5]['Time (s)']
            n_cycle = group[5]['Cycle No.']
            n_c = len(times)
            m_c = round(max(times),6)
            print ('from file: ', self.file_name[-7:], ', length of channel: ', n_c, 'cycle time (rounded):', m_c)
            self.total_counts[self.file_name[-7:]] += self.bin_data()
        #     #%%
        #     self.cycle_data = np.array(n_cycle[:])
        #     self.No_cycles = max(n_cycle)
        #     if self.cycle_max == 'max':                        
        #         self.data = times[:]
        #         print ('No cycles is:', self.No_cycles, 'length of time list:', len(self.data))
        #     else:
        #         top_cycle = self.cycle_data.tolist().index(self.cycle_max)
        #         self.data = times[:top_cycle]
        #         print ('top_cycle index is:', top_cycle, 'length of time list:', len(self.data))    
            
        #     self.total_counts[]
        # print (' Handling took : ' + str(round(time() - self.atime,4)) + ' s ')

    def tdms_handler_photons_current(self,skips):
        self.atime = time()
        cycle_skip = skips
        print ('skips: ', skips)
        if os.path.exists(self.tdms_path):
            with TdmsFile.open(self.tdms_path) as tdms_file:
                group= tdms_file.groups()
                yields = []
                parsback_values = []
                count_rate = []
        #     # if gr.properties['Description'] == self.channel_names[self.No_Chan[0]]: # the name of the channel might change
                # print (group[1].channels())
                times = group[1]['Time (s)']
                cycles = group[1]['Cycle No.']
                wls = group[1]['Laser Frequency (THz)']
                final_pows = group[1]['Power (mW)']
                # print (wls[20:200])
                # print (final_pows[20:200])
                n_c = len(times)
                m_c = round(max(times),6)
                
                print ('from file: ', self.file_name[-7:], ', length of channel: ', 1 ,n_c, 'cycle time (rounded):', m_c)
                self.No_cycles = int(max(cycles))
                cycles_np = cycles[:]
                cycle_0 = int(cycles[0])
                
                if self.cycle_max == 'none':  
                    top_cycle = self.No_cycles                      
                    print ('No cycles is:', top_cycle)
                else:
                    top_cycle = self.cycle_max
                    print ('The max set cycle is: ', top_cycle)

                No = 10
                delta_times = (self.stop - self.start)/No
                pow_matrix = np.zeros((top_cycle,No))
                wl_matrix = np.zeros((top_cycle,No))
                while cycle_0 < top_cycle:
                    m0 = cycles_np.tolist().index(cycle_0)
                    # print ('m0 in ', cycle_0, m0)
                    m1 = cycles_np.tolist().index(cycle_0 + 1)
                    # print ('m1 in ', cycle_0, m1)   
                    elem0 = 0
                    ini = self.start
                    for i in range(No):
                        elem1 = fn.find_nearest(times[m0:m1],ini)[0]
                        if len(final_pows[m0 + elem0:m0 + elem1]) > 0:
                            pow_matrix[cycle_0][i] = np.average(final_pows[m0 + elem0:m0 + elem1])
                            wl_matrix[cycle_0][i] = np.average(wls[m0 + elem0:m0 + elem1])
                        elem0 = elem1
                        ini += delta_times
                    cycle_0 += 1
                    print (elem0, elem1)
                fig1, axs1 = plt.subplots()
                l1 = axs1.imshow(pow_matrix,aspect='auto',origin='lower',cmap='flag',vmin=np.min(pow_matrix),vmax=np.max(pow_matrix))
                divider1 = make_axes_locatable(axs1)
                cax1 = divider1.append_axes("right", size="5%", pad=0.05)
                fig1.colorbar(l1, cax=cax1)
                fig, axs = plt.subplots()
                l2 = axs.imshow(wl_matrix,aspect='auto',origin='lower',cmap='flag',vmin=np.min(wl_matrix),vmax=np.max(wl_matrix))
                divider2 = make_axes_locatable(axs)
                axs.set_title('wavelengths')
                cax2 = divider2.append_axes("right", size="5%", pad=0.05)
                fig.colorbar(l2, cax=cax2)               

#%%
                    
                    # if cycles[m0] not in cycle_skip:
                    #     for k in range(m1 - m0):
                    #         wls_all[lams].append(times[m0 + k - 1])
                    # lams += 1
                    # # print (str(next_cycle) + ' took : ' + str(round(time() - atime,4)) + ' s ')
                    # cycle_0 += 1
                    # if lams == len(wavelengths):
                    #     lams = 1
                
    #             if self.cycle_max == 'none':                        
    #                 self.data = channel[:]
    #                 print ('No cycles is:', self.No_cycles, 'length of time list:', len(self.data))
    #             else:
    #                 top_cycle = self.cycle_data.tolist().index(self.cycle_max)
    #                 self.data = channel[:top_cycle]
    #                 print ('top_cycle index is:', top_cycle, 'length of time list:', len(self.data))

    #             yields.append(self.bin_data())
    #             parsback_values.append([self.file_name, self.No_cycles, n_c,m_c,0,self.x_mid,self.l,self.start,self.stop]) # file_name, length of channel, #No.cycles,storage_time,wl,firstpointtim
    #         if self.iscountsOn:
    #             count_rate.append(self.countRate())
    #             self.save_data(yields, parsback_values,count_rate,0)
    #         elif not self.iscountsOn:
    #             count_rate = 0
    #             self.save_data(yields, parsback_values,count_rate,0)
    #     else:
    #         print ('file: ', self.tdms_path, ' does not exist in that folder')
                    
    #     print (' Handling took : ' + str(round(time() - self.atime,4)) + ' s ')
    
    # def hdf5_handler(self):
    #     atime = time()
    #     yields = []
    #     parsback_values = []
    #     count_rate = []
    #     savePath = self.save_path + self.file_name + '.hdf'
    #     with h5py.File(savePath,'r') as f: 
    #         dset = f[str(self.hdf5_groupname)]
    #         for l,i in enumerate(self.channels):
    #     #     # if gr.properties['Description'] == self.channel_names[self.No_Chan[0]]: # the name of the channel might change
    #             channel = dset['Channel '+ str(i)]['Time (s)']
    #             n_cycle = dset['Channel '+ str(i)]['Cycle No.']
    #             self.cycle_data = n_cycle[:]
    #             print (type(n_cycle),type(self.cycle_data))

    #             self.No_cycles = max(n_cycle)
    #             if self.cycle_max == 'none':                        
    #                 self.data = channel[:]
    #                 print ('No cycles is:', self.No_cycles, 'length of time list:', len(self.data))
    #             else:
    #                 top_cycle = self.cycle_data.tolist().index(self.cycle_max)
    #                 self.data = channel[:top_cycle]
    #                 print ('top_cycle index is:', top_cycle, 'length of time list:', len(self.data))
    #             n_c = len(self.data)
    #             m_c = round(max(self.data),6)
    #             print ('from file: ', self.file_name[-4:], ', length of channel: ', i ,n_c, 'cycle time (rounded):', m_c)

    #             yields.append(self.bin_data())
    #             parsback_values.append([self.file_name, self.No_cycles, n_c,m_c,0,self.x_mid,self.l,self.start,self.stop]) # file_name, length of channel, #No.cycles,storage_time,wl,firstpointtim
    #         if self.iscountsOn:
    #             count_rate.append(self.countRate())
    #             self.save_data(yields, parsback_values,count_rate,0)
    #         elif not self.iscountsOn:
    #             count_rate = 0
    #             self.save_data(yields, parsback_values,count_rate,0)
    #         print (' Handling took : ' + str(round(time() - atime,4)) + ' s ')
                    
    def hdf5_wavelengths(self,wavelengths):
        atime = time()
        self.laser_on = True
        savePath = self.save_path + self.file_name + '.hdf'
        f = h5py.File(savePath,'r')
        dset = f[str(self.hdf5_groupname)]
        for l,i in enumerate(self.channels):
            times_ch = dset['Channel '+ str(i)]['Time (s)']
            cycles_ch = dset['Channel '+ str(i)]['Cycle No.']
            wls_ch = dset['Channel '+ str(i)]['Wavelength (nm)']
            wls = wls_ch[:]
            times = times_ch[:]
            cycles = cycles_ch[:]
            n_c = len(times)
            m_c = round(max(times),6)
            self.No_cycles = max(cycles)
            yields = []
            parsback_values = []
            # print ('the length of wls is: ', len(wls), 'first 10 values : ', wls[:10])
            if self.cycle_max == 'none':                        
                print ('No cycles is:', self.No_cycles, 'length of time list:', len(times))
            else:
                top_cycle = self.cycle_data.tolist().index(self.cycle_max)
                times = times[:top_cycle]
                wls= wls[:top_cycle]
                print ('top_cycle index is:', top_cycle, 'length of time list:', len(self.data))
                
            for wn, wv in enumerate(wavelengths):
                w_list = np.where(wls == wv)[0]
                data = []
                print ('length of wl list with wl ', wv,' nm: ', len(w_list), '\n') 
                if len(w_list) == 0:
                    print ('file ', self.file_name, ' did not have any wavelength,', wv)
                elif len(w_list) != 0:  
                    next_cycle = cycles[w_list[0]]
                    cycle_counter = 1
                    for val in w_list:
                        data.append(times[val])
                        ''' i needed the arbitrary time because of the time it takes the laser to change to the next wavelength
                        we should give somewhere around 1 second 
                        '''
                        if next_cycle != cycles[val]:
                            next_cycle = cycles[val]
                            cycle_counter += 1
                    self.data = data  
                    print ('wv' + str(wv) +' took : ' + str(round(time() - atime,4)) + ' s ')
                    yields.append(self.bin_data())
                    # power = self.energy_wl(_file, wavelength)
                    parsback_values.append([self.file_name, self.No_cycles,n_c, m_c,wv,cycle_counter,self.x_mid])
           
            if len(w_list) != 0:
                self.save_data(yields, parsback_values,0,wavelengths)
            else:
                print ('there is no values to save :( ')
            self.counter += 1 # this is for associating each channel 
        # self.save_data(yields, parsback_values,0)
        print ('hdf5 handling' + ' took : ' + str(round(time() - atime,4)) + ' s ')
    
    def tdms_wavelengths(self, wavelength,_file, arb_time):
        # self.wavelength = wavelength
        self.file_name = _file
        self.laser_on = True
        parsback_values = []
        tdms_path = self.DES_path +  _file + '.tdms'
        yields = []
        count_rate = []
        atime = time()
        with TdmsFile.open(tdms_path) as tdms_file:
            group= tdms_file.groups()
            for l,i in enumerate(self.channels):
                wls_ = group[i]['Wavelength (nm)']
                wls = wls_[:]
                wls = np.array(wls)
                # print (wls[0:10], wls[-50:-1])
                w_list = np.where(wls == wavelength)[0]
                print ('length of wl list with wl ', wavelength,' nm: ', len(w_list), '\n') 
                data = []
                cycle_counter = 1

                if len(w_list) == 0:
                    print ('file ', _file, ' did not have any wavelength,', wavelength)
                    break
                g = group[i]['Time (s)']
                channeldata = g[:]
                gcycle = group[i]['Cycle No.']
                cycledata = gcycle[:]
                inx = len(cycledata)
                self.No_cycles = max(cycledata)
                print ('length of cycledata: ', inx)
                m_c = max(channeldata)
                print ('cycle length: ' , round(m_c,4))
                next_cycle = cycledata[w_list[0]]
                for val in w_list:
                    data.append(channeldata[val])
                    ''' i needed the arbitrary time because of the time it takes the laser to change to the next wavelength
                    we should give somewhere around 1 second 
                    '''
                    if next_cycle != cycledata[val] and l == 0 and channeldata[val] > arb_time:
                        next_cycle = cycledata[val]
                        cycle_counter += 1
                print ('cycle_counter: ', cycle_counter)
                power = self.energy_wl(_file, wavelength)
                parsback_values.append([_file, round(m_c,2),cycle_counter, inx, round(power,2), wavelength])
                z = self.bin_data(data)  
                yields.append(z)
  
                # parsback values contain: wavelength, max time, No. cycles, cycles per wavelength, backgnd, std backgnd, start, stop 
        if len(w_list) != 0:
            self.save_data(yields, parsback_values,count_rate,wavelength)
        else:
            print ('there is no values to save :( ')
        print (str(wavelength) + ' took : ' + str(round(time() - atime,4)) + ' s ')
    
        return np.array(yields)

    def hdf5_wavelengths_Gustav(self,wavelengths,skips):
        self.atime = time()
        cycle_skip = skips
        self.laser_on = True
        f = h5py.File(self.savePath,'r')
        dset = f[str(self.hdf5_groupname)]
        wls_all = [[] for h in range(len(wavelengths))]
        for l,i in enumerate(self.channels):
            times_ch = dset['Channel '+ str(i)]['Time (s)']
            cycles_ch = dset['Channel '+ str(i)]['Cycle No.']
            wls_ch = np.array(dset['Channel '+ str(i)]['Wavelength (nm)'])
            # w_list = np.where(wls == wavelength)[0]
            # wls = wls_ch[:]
            times = times_ch[:]
            cycles = cycles_ch[:]
            m_c = round(max(times),6)
            self.No_cycles = max(cycles)
            yields = []
            parsback_values = []

            # print ('the length of wls is: ', len(wls), 'first 10 values : ', wls[:10])
            if self.cycle_max == 'none':                        
                print ('No cycles is:', self.No_cycles, 'length of time list:', len(times))
                top_cycle = self.No_cycles
            else:
                top_cycle = self.cycle_max
                # times = times[:top_cycle]
                # wls= wls[:top_cycle]
            next_cycle = cycles[0]
            print ('top cycle', top_cycle)
            print ('initial cycle:', next_cycle)
            lams = 0
            while next_cycle < top_cycle:
                m0 = cycles.tolist().index(next_cycle)
                # print ('m0 in ', next_cycle, m0)
                m1 = cycles.tolist().index(next_cycle + 1)
                # print ('m1 in ', next_cycle, m1)
                if cycles[m0] not in cycle_skip:
                    for k in range(m1 - m0):
                        wls_all[lams].append(times[m0 + k - 1])
                lams += 1
                # print (str(next_cycle) + ' took : ' + str(round(time() - atime,4)) + ' s ')
                next_cycle += 1
                if lams == len(wavelengths):
                    lams = 1
                    
            print ('i retrieved all the hits for all wavelengths, proceeding with binning')
            yields = []
            parsback_values = []
            print ('shape of yields', np.shape(wls_all[0]))
            print (wls_all[0][0])

            for j in range(len(wavelengths)):
                n_c = len(wls_all[j])
                self.data = wls_all[j]
                yields.append(self.bin_data())
                parsback_values.append([self.file_name, n_c, m_c, self.x_mid, wavelengths[j]])
            print ('shape of yields', np.shape(yields))
            self.save_pickle_data(yields, parsback_values,0,wavelengths)

    def hdf5_wavelengths_LPTS(self,wavelengths,cycle_slice,skips):
        self.atime = time()
        self.laser_on = True
        f = h5py.File(self.savePath,'r')
        dset = f[str(self.hdf5_groupname)]
        for l,i in enumerate(self.channels):
            times_ch = dset['Channel '+ str(i)]['Time (s)']
            cycles_ch = dset['Channel '+ str(i)]['Cycle No.']
            wls_ch = np.array(dset['Channel '+ str(i)]['Wavelength (nm)'])
            fluences = np.array(dset['Channel '+ str(i)]['Energy (mJ)'])
            # data_wls = [[] for h in range(len(wavelengths))]
            # data_final_pows = [[] for h in range(len(wavelengths))]
            yields = list()
            parsback_values = list()
            counts_per_wv = list()
            for j,wv in enumerate(wavelengths):
                data_wls = list()
                data_final_pows = list()
                w_list = np.where(wls_ch == wv)[0]
                # print ('length of wl list with wl ', wv,' nm: ', len(w_list), '\n') 
                if len(w_list) == 0:
                    print ('file ', self.file_name, ' did not have any wavelength,', wv)
                elif len(w_list) != 0:  
                    next_cycle = cycles_ch[w_list[0]]
                    cycle_counter = 1
                    for val in w_list:
                        data_wls.append(times_ch[val])
                        if fluences[val] > 20*0.9:
                            data_final_pows.append(fluences[val])
                        if next_cycle != cycles_ch[val]:
                            next_cycle = cycles_ch[val]
                            cycle_counter += 1
                print ('wv' + str(wv) +' took : ' + str(round(time() - self.atime,4)) + ' s ')
                max_wv_length = len(data_wls)
                counts_per_wv.append(max_wv_length)
                print ('wv' + str(wv) + ' is : ' + str(max_wv_length))
                self.data = data_wls
                self.final_pows = data_final_pows
                yields.append(self.bin_data())
                # parsback_values.append([self.file_name, max_wv_length, wv,0,0])
                parsback_values.append([self.file_name, max_wv_length, wv, np.average(data_final_pows),np.std(data_final_pows)])
            self.save_pickle_data(yields, parsback_values,0,wavelengths)
            plt.figure(l)
            plt.scatter(wavelengths,counts_per_wv)

    def hdf5_wavelengths_LPTS_2(self,wavelengths,skips,step_wv,ini_final_pows):
        self.atime = time()
        self.laser_on = True
        cycle_skip = skips[0]
        wls = wavelengths
        ini_final_pows = ini_final_pows
        f = h5py.File(self.savePath,'r')
        dset = f[str(self.hdf5_groupname)]
        final_wls = list()
        final_pows = list()
        counts_per_wv = [[] for h in range(len(self.channels))]

        for l,i in enumerate(self.channels):
            times_ch = dset['Channel '+ str(i)]['Time (s)']
            cycles_ch = dset['Channel '+ str(i)]['Cycle No.']
            wls_ch = np.array(dset['Channel '+ str(i)]['Wavelength (nm)'])
            fluences = np.array(dset['Channel '+ str(i)]['Energy (mJ)'])
            cycles = cycles_ch[:]
            times = times_ch[:]
            wavis = wls_ch[:]
            powis = fluences[:]
            ini_power = max(powis)
            self.No_cycles = max(cycles_ch)
            all_columns = list()
            if self.cycle_max == 'max':                        
                top_cycle = self.No_cycles
            else:
                top_cycle = self.cycle_max
            print ('No cycles is:', self.No_cycles, 'length of time list:', len(times_ch), 'top cycle: ', top_cycle)
            next_cycle = np.arange(cycle_skip,top_cycle,1)
            first_cycle = cycles[0]
            first_cycle_pos = cycles.tolist().index(first_cycle)
            first_wv = wavis[first_cycle_pos]
            print ('first wavelength: ', first_wv, ' max laser power: ', ini_power)
            for cn, c in enumerate(next_cycle):
                if l == 0:
                    if first_wv > max(wavelengths):
                        first_wv = min(wavelengths)
                    # final_wls.append(first_wv)
                    elem_w = wavelengths.tolist().index(first_wv)
                    if c % step_wv == 0:
                        first_wv += 1
                        
                if c in cycles:
                    first_cycle += 1
                    time_data = list()
                    final_pows_data = list()
                    c_list = np.where(cycles == c)[0]
                    final_wls.append(wavis[c_list[0]])
                    for elem in c_list:
                        time_data.append(times[elem])
                        if l == 0:
                            if powis[elem] > ini_power*0.8:
                                final_pows_data.append(powis[elem])
                    self.data = time_data
                    counts_per_wv[l].append(self.bin_data_LPTS()[0])
                    if l == 0:
                        if len(final_pows_data) != 0:
                            final_pows.append(np.average(final_pows_data))
                        else:
                            final_pows.append(ini_final_pows[elem_w])
                else:
                    counts_per_wv[l].append(0)
                    if l == 0:
                        final_pows.append(ini_final_pows[elem_w])
                        if c % step_wv == 0 and len(final_wls) != 0:
                            final_wls.append(final_wls[-1] + 1)
                        elif c % step_wv == 0 and len(final_wls) == 0:
                            final_wls.append(final_wls[-1])

            print ('channel' + str(self.channels[l]) +' took : ' + str(round(time() - self.atime,4)) + ' s ')

            if l == 0:
                fig,ax = plt.subplots()
                ax.plot(next_cycle,counts_per_wv[l], alpha = 0.5, color = 'orange')
                ax.plot(next_cycle,np.array(final_pows)/10, alpha = 0.5, color = 'blue')
                ax2 = ax.twinx()
                ax2.plot(next_cycle,final_wls,alpha = 0.3,color = 'black')
                
        print (first_cycle, len(next_cycle), len(counts_per_wv[0]),len(final_wls), len(final_pows))
        for j in range(len(next_cycle)):
            all_columns.append([next_cycle[j],counts_per_wv[0][j],counts_per_wv[1][j],final_wls[j],final_pows[j]])
        self.save_data_LPTS(np.array(all_columns))

    def hdf5_LPTS(self,wls,step_wv):
        self.atime = time()
        self.set_data_LPTS()
        self.laser_on = True
        final_wls = list()
        final_pows = list()
        middle_pows = list()
        all_columns = list()
        counts_per_wv = [[] for h in range(len(self.channels))]
        if self.cycle_max == 'max':
            top_cycle = self.No_cycles
        else:
            top_cycle = self.cycle_max
        print ('No cycles is:', self.No_cycles, 'length of time list:', len(self.ts[1]), 'top cycle: ', top_cycle)
        next_cycle = np.arange(self.cycle_min,top_cycle,1)
        first_cycle = self.c4[0]
        first_wv = self.wavis[0]
        self.first_scan = False
        
        for wi in wls:
            wavis_list = np.where(self.wavis == wi)[0]
            powis_list = list()
            for w in wavis_list:
                if self.powis[w] > 0.7*self.max_pow:
                    powis_list.append(self.powis[w])
            final_pows.append([wi,np.average(powis_list)])     
            
        for cn, c in enumerate(next_cycle):
            self.acuired_wavelength = False
            self.aquired_pow = False
            if c in self.c3:
                time_per_wv = [[] for i in range(2)]
                c_list = np.where(self.c3 == c)[0] 
                for i in range(2):
                    for elem in c_list:
                        time_per_wv[i].append(self.ts[i][elem])
                        if not self.acuired_wavelength and self.ts[1][elem] > 1:
                            final_wls.append(self.wavis[elem])
                            self.acuired_wavelength = True
                        if not self.aquired_pow and self.ts[1][elem] > 1:
                            middle_pows.append(self.powis[elem])
                            self.aquired_pow = True
                    counts_per_wv[i].append(self.bin_data_LPTS(time_per_wv[i])[0])
                    # if c == 1197:
                    #     print (c, ' in c3')
                        
            else:
                time_per_wv = list()
                counts_per_wv[0].append(0)
                c_list = np.where(self.c4 == c)[0]
                for en, elem in enumerate(c_list):
                    time_per_wv.append(self.ts[1][elem]) # dont forget, 0 is for signal, 1 background
                    if not self.acuired_wavelength and self.ts[1][elem] > 1:
                        final_wls.append(self.wavis[elem])
                        self.acuired_wavelength = True
                    if not self.aquired_pow and self.ts[1][elem] > 1:
                        middle_pows.append(self.powis[elem])
                        self.aquired_pow = True
                counts_per_wv[1].append(self.bin_data_LPTS(time_per_wv)[0])
                
            if first_wv > max(wls):
                first_wv = min(wls)
            if not self.aquired_pow:
                wl_position = wls.tolist().index(first_wv)
                middle_pows.append(final_pows[wl_position][1])
                self.aquired_pow = True
            if not self.acuired_wavelength:
                final_wls.append(first_wv)
                self.acuired_wavelength = True
            if c % step_wv == 0:
                first_wv += 1

        fig,ax = plt.subplots()
        ax.plot(next_cycle,counts_per_wv[0], alpha = 0.5, color = 'orange')
        ax.plot(next_cycle,counts_per_wv[1], alpha = 0.5, color = 'black')
        ax2 = ax.twinx()
        ax2.plot(next_cycle,final_wls,alpha = 0.3,color = 'green')        
        ax.plot(next_cycle,np.array(middle_pows)/10, alpha = 0.5, color = 'blue')
        
        print (first_cycle, len(next_cycle), len(counts_per_wv[0]),len(final_wls), len(final_pows))
        for j in range(len(next_cycle)):
            all_columns.append([next_cycle[j],counts_per_wv[0][j],counts_per_wv[1][j],final_wls[j],middle_pows[j]])
        self.save_data_LPTS(np.array(all_columns),final_pows)
    
    def hdf5_LPTS_2(self,wls,b_fac,b_range):
        hc = 1239.84193 #eV/nm, this is the factor to convert nm -> eV
        self.atime = time()
        self.set_data_LPTS()
        self.laser_on = True
        final_pows = list()
        self.counts_G = list()
        self.counts_B = list()
        cycles_all = list()
        counts_per_wv = np.zeros(len(wls))
        bi = b_range
        bf = b_range + 7
        b_fac = b_fac
        if self.cycle_max == 'max':
            top_cycle = self.No_cycles
        else:
            top_cycle = self.cycle_max
        print ('file: ', self.file_name)
        print ('No cycles is:', self.No_cycles, 'length of time list:', len(self.ts[1]), 'top cycle: ', top_cycle)
        print ('max time: ', self.max_time)

        next_cycle = np.arange(self.cycle_min,top_cycle,1)
        next_cycle = next_cycle.tolist()
        for i in range(len(self.skips)):
            if self.skips[i] in next_cycle:
                next_cycle.remove(self.skips[i])
        self.first_cycle = self.c3[0]
        self.raw_dG = list()
        self.raw_dB = self.bin_data_quick(self.ts[1],100,0.1,220)

        for wk, wi in enumerate(wls):
            wavis_list_3 = np.where(self.wavis3 == wi)[0]
            wavis_list_4 = np.where(self.wavis4 == wi)[0]
            powis_list = list()
            time_per_wv = list()
            back_list = list()
            for w in wavis_list_3:
                powis_list.append(self.powis[w])
            max_pow_wl = max(powis_list)
            powis_list_2 = list()

            for w in wavis_list_3:
                time_per_wv.append(self.ts[0][w])
                if self.powis[w] > 0.1*max_pow_wl: 
                    powis_list_2.append(self.powis[w])

            for w in wavis_list_4:
                back_list.append(self.ts[1][w])

            counts_per_wv[wk] = self.bin_data_LPTS(time_per_wv,self.start,self.stop)[0]
            # self.total_counts_per_wavelength[wk] += self.bin_data_LPTS(time_per_wv,self.start,self.stop)[0]
            self.total_counts[str(wi)][0] += self.bin_data_LPTS(time_per_wv,self.start,self.stop)[0]
            self.total_counts[str(wi)][2] += self.bin_data_LPTS(back_list,self.start,self.stop)[0]/b_fac
            self.raw_dG.append(self.bin_data_quick(time_per_wv,100,0.1,220))

            if len(powis_list) != 1:
                final_pows.append(np.average(powis_list_2)) 
                self.total_counts[str(wi)][1] = max_pow_wl
                # print ('This is the average of the power in ', wi, ' : ', np.average(powis_list_2))
            else:
                print ('i could not find a single power value, so i take the max ', wi, ' : ', self.max_pow)
                self.total_counts[str(wi)][1] = max_pow_wl
                final_pows.append(self.max_pow)

        # self.bin_data_LPTS_delayed_signal(delayed_signal,)
        self.norm_pows = np.array(final_pows)/max(final_pows)
        self.darkcounts = self.bin_data_LPTS(self.ts[0],bi,bf)[0]
        while self.first_cycle < next_cycle[-1] + 1:
            cycles_all.append(self.first_cycle)
            if self.first_cycle in self.c3:
                elem0 = self.c3.tolist().index(self.first_cycle)
                elem10 = self.c4.tolist().index(self.first_cycle)
                cd = 1
                cycle_is_there = False
                while not cycle_is_there:
                    if self.first_cycle + cd in self.c3:
                        # print ('now in : ' + str(self.first_cycle + cd))
                        elem1 = self.c3.tolist().index(self.first_cycle+cd)
                        elem11 = self.c4.tolist().index(self.first_cycle+cd)
                        cycle_is_there = True
                    else:
                        cd += 1
                    # print (elem0,elem1)
                self.counts_G.append(self.bin_data_LPTS(self.ts[0][elem0:elem1],self.start,self.stop))
                self.counts_B.append(self.bin_data_LPTS(self.ts[1][elem10:elem11],self.start,self.stop))
                self.first_cycle += cd
            else:
                self.counts_G.append(0)
                self.counts_B.append(0)
                self.first_cycle += 1
        
        if self.counter < 2:
            plt.figure(self.l + 90)
            # plt.plot(self.raw_dG.T[0][:],self.raw_dG.T[0][:], color = 'gold', label = 'Gated')
            plt.plot(cycles_all,np.array(self.counts_B)*5, color = 'blue', label = 'Back x 20')
            plt.plot(cycles_all,self.counts_G - np.array(self.counts_B)/b_fac, color = 'orange', label = 'difference')
            plt.plot(cycles_all,self.counts_G, color = 'green', label = 'total G')

            # plt.ylabel('Counts')
            plt.ylabel('Counts')
            plt.yscale('linear')
            plt.xlabel('Cycle No.')
            plt.legend(loc = 'best')
            # plt.title(self.file_name)
            # plt.savefig(self.save_path + '{}_count_rates'.format(self.dia) + '.pdf', dpi = 300, bbox_inches='tight')
            
        
        # back_per_cycle_per_second = backgnd/(20*self.No_cycles*(bf - bi))
        # err_back = np.sqrt(backgnd)/(20*self.No_cycles*(bf - bi))
        # xw = hc/wls
        # total_meas_time = (self.stop - self.start)*self.No_cycles//len(wls)
        # yw = (np.array(counts_per_wv) - back_per_cycle_per_second*total_meas_time)/self.norm_pows
        # errw = np.sqrt(np.array(counts_per_wv))/self.norm_pows
        # print ('background values: ', back_per_cycle_per_second, err_back)
        # print ('real counts: ', yw[0],errw[0])
        # plt.figure()
        # plt.errorbar(xw,yw,yerr=errw,
        #               fmt ='.k', capthick=1.5,capsize=3,elinewidth=1.5, label = 'exp data ' + self.file_name)
        # plt.legend(loc = 'best')
        # self.save_data_LPTS(np.array(all_columns),final_pows)
    
    def hdf5_pumpingStates(self,wls):
        self.set_data_pumpingStates()
        self.atime = time()
        elem0,value0 = fn.find_nearest(self.nus,wls[0])
        elemf,valuef = fn.find_nearest(self.nus,wls[-1])
        # print (len(self.ts))
    
    def set_data_LPTS(self):
        f = h5py.File(self.savePath,'r')
        dset = f[str(self.hdf5_groupname)]
        times_ch4 = dset['Channel ' + str(self.channels[1])]['Time (s)']
        times_ch3 = dset['Channel ' + str(self.channels[0])]['Time (s)']
        times_chID = dset['Channel ' + str(self.channels[2])]['Time (s)']

        cycles_ch4 = dset['Channel '+ str(self.channels[1])]['Cycle No.']
        cycles_ch3 = dset['Channel '+ str(self.channels[0])]['Cycle No.']
        self.c4 = cycles_ch4[:]
        self.c3 = cycles_ch3[:]
        self.ts = np.array([times_ch3[:],times_ch4[:]], 'object')
        self.t_ID = times_chID[:]
        
        wls_ch3 = dset['Channel ' + str(self.channels[0])]['Wavelength (nm)']
        wls_ch4 = dset['Channel ' + str(self.channels[1])]['Wavelength (nm)']

        pulseEs_ch4 = dset['Channel ' + str(self.channels[0])]['Energy (mJ)']
        self.wavis3 = np.array(wls_ch3)
        self.wavis4 = np.array(wls_ch4)

        self.powis = np.array(pulseEs_ch4)
        self.max_pow = max(self.powis)
        self.No_cycles = max(cycles_ch3)
        self.max_time = self.ts[1][-1]
        print (len(self.ts[0]),len(self.ts[1]),len(np.array(wls_ch4)),len(np.array(wls_ch3)))
    
    def set_data_pumpingStates(self):
        with TdmsFile.open(self.tdms_path) as tdms_file:
            self.tdms_properties = tdms_file.properties
            self.group= tdms_file.groups()
            for l,i in enumerate(self.channels):
                try:
                    self.times_PMT = self.group[i]['Time (s)']
                    self.wls_PMT = self.group[i]['Laser Frequency (THz)']
                    self.cycles_PMT = self.group[i]['Cycle No.']
                    self.nus = self.wls_PMT[:]
                    self.ts = self.times_PMT[:]
                    self.cs = self.cycles_PMT[:]
                    self.max_cycle = max(self.cs)
                except Exception as e:
                    print (e)

        
    def hdf5_LPTS_3(self):
        f = h5py.File(self.savePath,'r')
        dset = f[str(self.hdf5_groupname)]
        wls_ch3 = dset['Channel ' + str(self.channels[0])]['Wavelength (nm)']
        self.wavelengths = list()
        wavis_0 = np.array(wls_ch3)
        wl0 = wavis_0[0]
        self.wavelengths.append(wl0)
        for wl in wavis_0:
            if wl != wl0:
                self.wavelengths.append(wl)
                wl0 = wl
                
        workbook = openpyxl.load_workbook(self.save_path + 'wider_wavelength_scan.xlsx')
        worksheet = workbook['total_cycles_per_wavelength4']
        
        worksheet.cell(row=1, column=1).value = 'wavelengths (nm)'

        for i, item in enumerate(self.wavelengths):
            worksheet.cell(row=i+2, column=1).value = item
        
        workbook.save(self.save_path + 'wider_wavelength_scan.xlsx')

    
    def hdf5_wavelengths_Gustav_gated(self,wavelengths,skips,gate_times):
        atime = time()
        cycle_skip = skips
        gates = gate_times
        self.laser_on = True
        savePath = self.save_path + self.file_name + '.hdf'
        f = h5py.File(savePath,'r')
        dset = f[str(self.hdf5_groupname)]
        wls_all = [[] for h in range(len(wavelengths))]
        for l,i in enumerate(self.channels):
            times_ch = dset['Channel '+ str(i)]['Time (s)']
            cycles_ch = dset['Channel '+ str(i)]['Cycle No.']
            # wls_ch = dset['Channel '+ '2']['Wavelength (nm)']
            # wls = wls_ch[:]
            times = times_ch[:]
            cycles = cycles_ch[:]
            m_c = round(max(times),6)
            self.No_cycles = max(cycles)
            yields = []
            parsback_values = []

            # print ('the length of wls is: ', len(wls), 'first 10 values : ', wls[:10])
            if self.cycle_max == 'none':                        
                print ('No cycles is:', self.No_cycles, 'length of time list:', len(times))
                top_cycle = self.No_cycles
            else:
                top_cycle = self.cycle_max
                # times = times[:top_cycle]
                # wls= wls[:top_cycle]
            next_cycle = cycles[0]
            print ('top cycle', top_cycle)
            print ('initial cycle:', next_cycle)
            lams = 0
            while next_cycle < top_cycle:
                m0 = cycles.tolist().index(next_cycle)
                # print ('m0 in ', next_cycle, m0)
                m1 = cycles.tolist().index(next_cycle + 1)
                # print ('m1 in ', next_cycle, m1)
                if cycles[m0] not in cycle_skip:
                    for k in range(m1 - m0):
                        gated_counts = times[m0 + k - 1]
                        if gated_counts in gates:
                            wls_all[lams].append(gated_counts)
                lams += 1
                # print (str(next_cycle) + ' took : ' + str(round(time() - atime,4)) + ' s ')
                next_cycle += 1
                if lams == len(wavelengths):
                    lams = 1
                    
            print ('i retrieved all the hits for all wavelengths, proceeding with binning')
            yields = []
            parsback_values = []
            print ('shape of yields', np.shape(wls_all[0]))
            print (wls_all[0][0])

            for j in range(len(wavelengths)):
                n_c = len(wls_all[j])
                self.data = wls_all[j]
                yields.append(self.bin_data())
                parsback_values.append([self.file_name, n_c, m_c, self.x_mid, wavelengths[j]])
            print ('shape of yields', np.shape(yields))
            self.save_data(yields, parsback_values,0,wavelengths)
            
    def countRate(self):
        count= 0
        xy_cycle = []
        y_sts = []
        while count < self.No_cycles:
            c_list = np.where(self.cycle_data == count + 1)[0]
            yield_list = []
            for val in c_list:
                if self.start < self.data[val] < self.stop:
                    yield_list.append(self.data[val])
            sum_yi = 0
            for yi in yield_list:
                sum_yi += yi
            xy_cycle.append([count,sum_yi])
            y_sts.append(sum_yi)
            count += 1
        return np.array(xy_cycle)
        

    def bin_data(self):
        '''
        gives linear or log binning
        '''
        data_x_parts = []
        x_step = self.start
        if self.method == 'linear':
            his, bin_edges = np.histogram(self.data, bins = self.Nbins, range=(self.start, self.stop))
        elif self.method == 'log': 
            his, bin_edges = np.histogram(self.data, bins = np.logspace(np.log10(self.start),np.log10(self.stop), self.Nbins))
        x_middle = (bin_edges[1]-bin_edges[0])/2
        x_step += x_middle
        x_s0 = x_middle*2
        self.x_mid = x_middle 
        for k in range(len(his)):
            x_s = bin_edges[k+1]-bin_edges[k] # if the size of the step changes over bins, then this is actually very clever
            data_x_parts.append([x_step,his[k],np.sqrt(his[k]),x_s])
            x_step += x_s  
        return np.array(data_x_parts)

    def bin_data_quick(self,data,bins,start,stop):
        '''
        gives linear or log binning
        '''
        self.method = 'linear'
        data_x_parts = []
        x_step = start
        his, bin_edges = np.histogram(data, bins = bins, range=(start, stop))
        # elif self.method == 'log': 
            # his, bin_edges = np.histogram(self.data, bins = np.logspace(np.log10(self.start),np.log10(self.stop), self.Nbins))
        x_middle = (bin_edges[1]-bin_edges[0])/2
        x_step += x_middle
        x_s0 = x_middle*2
        self.x_mid = x_middle 
        for k in range(len(his)):
            x_s = bin_edges[k+1]-bin_edges[k] # if the size of the step changes over bins, then this is actually very clever
            data_x_parts.append([x_step,his[k],np.sqrt(his[k]),x_s])
            x_step += x_s  
        return np.array(data_x_parts)

    def bin_data_LPTS(self,data,start,stop):
        his, bin_edges = np.histogram(data, bins = 1, range=(start,stop))
        return his
    
    def bin_data_LPTS_delayed_signal(self,data,bins,start,stop):
        his, bin_edges = np.histogram(data, bins = 1, range=(start,stop))
        return his

    def wavelength_check(self, _file):
        tdms_path = self.DES_path + _file + '.tdms'
        with TdmsFile.open(tdms_path) as tdms_file:
            # print (tdms_file.properties)
            group= tdms_file.groups()
            # for gr in group:
                # print (gr.properties['Description'])
            # print (self.channels)
            wls_ = group[self.channels[0]]['Wavelength (nm)']
            n_cycle = group[self.channels[0]]['Cycle No.']
            n_c = n_cycle[:]
            n_c = np.array(n_c)
            wls = wls_[:]
            wls = np.array(wls)
            first_elem = wls[0]
            for k,i in enumerate(wls):
                if i != first_elem:
                    print (first_elem, 'second element: ', i , ' in position ' , k, 'cycle: ', n_c[k])
                    break
                
        
    def doppler_shift(self,lam0_,theta,co):
        ''' not  relativistic'''
        ion_vel = self.ion_velocity()
        c = self.light_speed
        lam0 = lam0_*1e-9
        nu_0 = c/lam0
        if co:
            nu_p = nu_0*(1 - ion_vel/c*np.cos(theta*np.pi/180))
        else:
            nu_p = nu_0*(1 + ion_vel/c*np.cos(theta*np.pi/180))            
        print ('nu: ' ,c/nu_p, 'nu0: ', c/nu_0)
            
    '''
    Saving function1
    '''
        
    def save_data(self, z_sub, backgnd,c_rate,wavelengths):
        savefp = self.save_path + '/' + self.file_name
        if not os.path.exists(savefp + '/pars/'):
            os.makedirs(savefp + '/pars/')
            
        if self.laser_on:
            save_csv_name = savefp  + '/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' +str(self.l)
            save_csv_pars = savefp + '/pars/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' + str(len(wavelengths)) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_' + str(len(wavelengths)) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))

            for i in range(len(wavelengths)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}.csv'.format(wavelengths[i]), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])

        else:
            save_csv_name = savefp  + '/' + self.method + '_' + str(self.file_name)
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
        
            for i in range(len(self.channels)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}_{}_countRate.csv'.format(self.channels[i],self.l), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])
        print ('saving' + str(wavelengths) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')

        with open(save_csv_name + '_{}_{}_pars.csv'.format(self.channels[i],self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)
            
            
    def save_data_OPO_pow(self, z_sub, backgnd):
        savefp = self.save_path + '/Qsw'
        save_csv_name = savefp  + '/' + self.method + '_' + str(self.file_name)

        if not os.path.exists(savefp):
            os.makedirs(savefp)
        
        for i in range(len(self.channels)):
            # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
            with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                writer = csv.writer(arch, delimiter = '\t')
                writer.writerows(z_sub[i])
            print ('file created in:', savefp, ' under the name ', self.method + '_' + str(self.file_name) + '_{}_{}.csv'.format(self.channels[i],self.l))
            
        with open(save_csv_name + '_{}_{}_pars.csv'.format(self.channels[i],self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)
            
        print ('saving' + str(self.file_name) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')

            
            
    def save_data_LPTS(self,z,pows):
        savefp = self.save_path + '/LPTS/'
        if not os.path.exists(savefp):
            os.makedirs(savefp)
        print ('The shape of your data: ', np.shape(z))
        with open(savefp + 'LPTS{}.csv'.format(self.file_name), 'w') as arch:
            writer = csv.writer(arch, delimiter = '\t')
            writer.writerows(z)
        with open(savefp + 'LPTS{}_pows.csv'.format(self.file_name), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(pows)
            
    def save_pickle_data(self, z_sub, backgnd,c_rate,wavelengths):
        savefp = self.save_path + '/' + self.file_name
        if not os.path.exists(savefp):
            os.makedirs(savefp)
            
        if not os.path.exists(savefp + '/pars/'):
            os.makedirs(savefp + '/pars/')
        
        # if not os.path.exists(savefp + 'countRates/'):
            # print('making folder: ' + savefp + 'countRates/')
            # os.makedirs(savefp + 'countRates/')
            
            
        if self.laser_on:
            save_csv_name = savefp  + '/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' +str(self.l)
            save_csv_pars = savefp + '/pars/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' + str(len(wavelengths)) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_' + str(len(wavelengths)) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
    
            for i in range(len(wavelengths)):
                for j in range(len(self.channels)):
                    # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                    with open(save_csv_name + '_{}_{}.pickle'.format(self.channels[j],wavelengths[i]), 'wb') as arch:
                        try:
                            writer = pickle.dump(z_sub[i],arch, protocol = pickle.HIGHEST_PROTOCOL)
                        except Exception as ex:
                            print ('Error during pickling: ', ex)
    
                    if self.iscountsOn:
                        with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                            writer = csv.writer(arch, delimiter = '\t')
                            writer.writerows(c_rate[i])
    
        else:
            save_csv_name = savefp  + self.method + str(self.file_name) + str(self.l)
            save_csv_pars = savefp + 'pars/' + self.method + str(self.file_name) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
        
            for i in range(len(self.channels)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])
        print ('saving' + str(wavelengths) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')
    
            
    
    
        with open(save_csv_pars + '{}.csv'.format(self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)


    def save_energies(self,energy_data,save_filepath):
        savefp = self.save_path + save_filepath
        save_csv_name = savefp + str(self.file_name) + 'energies_wavelengths'
        if not os.path.exists(savefp):
            os.makedirs(savefp)
        if not os.path.exists(save_csv_name + '.csv'):
            with open(save_csv_name + '.csv', 'w') as f:
                writer = csv.writer(f, delimiter = '\t')
                writer.writerows(energy_data)
            
    def save_rate(self, cycleRate,save_filepath):
        savefp = self.save_path + save_filepath
        save_csv_name = savefp +  str(self.file_name) + 'cycleRate'
        if not os.path.exists(savefp):
            os.makedirs(savefp)
        if not os.path.exists(save_csv_name + '.csv'):
            with open(save_csv_name + '.csv', 'w') as f:
                writer = csv.writer(f, delimiter = '\t')
                writer.writerows(cycleRate)




class Analysis_laser_induced_signal:
    LE_LEBC = 3.25 #m
    HE_HEBC = 6.71
    # the first for S and the second for A ring
    LEBC_ring = [7.24,9.59] #m
    HEBC_ring = [9.93,7.80]
    ring_circ = [8.68,8.71] # m
    rad = ring_circ[0]/(2*np.pi) #m
    cen_RAES = 0
    light_speed = 299792458
    
    def __init__(self, source_path, saving_path, integrated_shots):
        # self.q = queue.Queue()    
        self.DES_path = source_path
        self.save_path = saving_path
        self.laser_on = False
        self.integrated_shots = integrated_shots
        if not os.path.exists(self.save_path):
            os.makedirs(self.save_path)
            print("Directory " , self.save_path,  " Created ")
        else:    
            print("Directory " , self.save_path ,  " already exists")
        self.channel_names = []
        self.channels = []


    def ini_params(self,exp, _file,start,stop,hdf5_groupname, n1,n2,bins,dump):
        self.file_name = _file
        self.experiment = exp
        self.bins = bins
        self.start = start
        self.stop = stop
        self.cycle_max = n2
        self.cycle_min = n1
        self.beam_dump = dump
        self.savePath = self.save_path + self.experiment + self.file_name + '.hdf'
        self.tdms_path = self.DES_path + self.experiment + self.file_name + '.tdms'
        # self.q.put(self.tdms_path)
        self.Signal_vs_power()

      
    def Signal_vs_power(self):
        self.atime = time()
        with TdmsFile.open(self.tdms_path) as tdms_file:
            # print (tdms_file.properties)
            group= tdms_file.groups()
        #     # if gr.properties['Description'] == self.channel_names[self.No_Chan[0]]: # the name of the channel might change
            times = group[5]['Time (s)']
            n_cycle = group[5]['Cycle No.']
            n_c = len(times)
            m_c = round(max(times),6)
            number_cycles = len(n_cycle)
            print ('from file: ', self.file_name[-7:], ', length of channel: ', n_c, 'cycle time (rounded):', m_c)
            self.b_norm = self.bin_data(times,self.beam_dump,self.beam_dump + 7e-2)
            back = 0
            for i in range(len(self.start)):
                signal = self.bin_data(times,self.start[i],self.stop[i])
                back += self.bin_data(times,self.stop[i] ,self.stop[i] + 24e-2)
                self.integrated_shots[self.file_name[-7:]][0] += (signal - self.b_norm/70)               
                self.integrated_shots[self.file_name[-7:]][1] += np.sqrt((signal + self.b_norm/70))
            self.integrated_shots[self.file_name[-7:]][2] = back/number_cycles
            self.integrated_shots[self.file_name[-7:]][3] = np.sqrt(back)/number_cycles

        print (' Handling of : ', self.file_name[-7:], ' took : ' + str(round(time() - self.atime,4)) + ' s ')

        # print ('All took : ' + str(round(time() - self.atime,4)) + ' s ')

    def bin_data(self,data,start,stop):
        his, bin_edges = np.histogram(data, bins = self.bins, range=(start, stop))
        return his

        
    def doppler_shift(self,lam0_,theta,co):
        ''' not  relativistic'''
        ion_vel = self.ion_velocity()
        c = self.light_speed
        lam0 = lam0_*1e-9
        nu_0 = c/lam0
        h_eV = 4.135667696e-15
        if co:
            nu_p = nu_0*(1 - ion_vel/c*np.cos(theta*np.pi/180))
        else:
            nu_p = nu_0*(1 + ion_vel/c*np.cos(theta*np.pi/180))            
        print ('nu: ' ,c/nu_p, 'nu0: ', c/nu_0)
        return nu_p*h_eV
            
    '''
    Saving function1
    '''
        
    def save_data(self, z_sub, backgnd,c_rate,wavelengths):
        savefp = self.save_path + '/' + self.file_name
        if not os.path.exists(savefp + '/pars/'):
            os.makedirs(savefp + '/pars/')
            
        if self.laser_on:
            save_csv_name = savefp  + '/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' +str(self.l)
            save_csv_pars = savefp + '/pars/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' + str(len(wavelengths)) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_' + str(len(wavelengths)) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))

            for i in range(len(wavelengths)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}.csv'.format(wavelengths[i]), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])

        else:
            save_csv_name = savefp  + '/' + self.method + '_' + str(self.file_name)
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
        
            for i in range(len(self.channels)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}_{}_countRate.csv'.format(self.channels[i],self.l), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])
        print ('saving' + str(wavelengths) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')

        with open(save_csv_name + '_{}_{}_pars.csv'.format(self.channels[i],self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)
            
            
    def save_data_OPO_pow(self, z_sub, backgnd):
        savefp = self.save_path + '/Qsw'
        save_csv_name = savefp  + '/' + self.method + '_' + str(self.file_name)

        if not os.path.exists(savefp):
            os.makedirs(savefp)
        
        for i in range(len(self.channels)):
            # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
            with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                writer = csv.writer(arch, delimiter = '\t')
                writer.writerows(z_sub[i])
            print ('file created in:', savefp, ' under the name ', self.method + '_' + str(self.file_name) + '_{}_{}.csv'.format(self.channels[i],self.l))
            
        with open(save_csv_name + '_{}_{}_pars.csv'.format(self.channels[i],self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)
            
        print ('saving' + str(self.file_name) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')

            
            
    def save_data_LPTS(self,z,pows):
        savefp = self.save_path + '/LPTS/'
        if not os.path.exists(savefp):
            os.makedirs(savefp)
        print ('The shape of your data: ', np.shape(z))
        with open(savefp + 'LPTS{}.csv'.format(self.file_name), 'w') as arch:
            writer = csv.writer(arch, delimiter = '\t')
            writer.writerows(z)
        with open(savefp + 'LPTS{}_pows.csv'.format(self.file_name), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(pows)
            
    def save_pickle_data(self, z_sub, backgnd,c_rate,wavelengths):
        savefp = self.save_path + '/' + self.file_name
        if not os.path.exists(savefp):
            os.makedirs(savefp)
            
        if not os.path.exists(savefp + '/pars/'):
            os.makedirs(savefp + '/pars/')
        
        # if not os.path.exists(savefp + 'countRates/'):
            # print('making folder: ' + savefp + 'countRates/')
            # os.makedirs(savefp + 'countRates/')
            
            
        if self.laser_on:
            save_csv_name = savefp  + '/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' +str(self.l)
            save_csv_pars = savefp + '/pars/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' + str(len(wavelengths)) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_' + str(len(wavelengths)) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
    
            for i in range(len(wavelengths)):
                for j in range(len(self.channels)):
                    # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                    with open(save_csv_name + '_{}_{}.pickle'.format(self.channels[j],wavelengths[i]), 'wb') as arch:
                        try:
                            writer = pickle.dump(z_sub[i],arch, protocol = pickle.HIGHEST_PROTOCOL)
                        except Exception as ex:
                            print ('Error during pickling: ', ex)
    
                    if self.iscountsOn:
                        with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                            writer = csv.writer(arch, delimiter = '\t')
                            writer.writerows(c_rate[i])
    
        else:
            save_csv_name = savefp  + self.method + str(self.file_name) + str(self.l)
            save_csv_pars = savefp + 'pars/' + self.method + str(self.file_name) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
        
            for i in range(len(self.channels)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])
        print ('saving' + str(wavelengths) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')
    
            
    
    
        with open(save_csv_pars + '{}.csv'.format(self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)


    def save_energies(self,energy_data,save_filepath):
        savefp = self.save_path + save_filepath
        save_csv_name = savefp + str(self.file_name) + 'energies_wavelengths'
        if not os.path.exists(savefp):
            os.makedirs(savefp)
        if not os.path.exists(save_csv_name + '.csv'):
            with open(save_csv_name + '.csv', 'w') as f:
                writer = csv.writer(f, delimiter = '\t')
                writer.writerows(energy_data)
            
    def save_rate(self, cycleRate,save_filepath):
        savefp = self.save_path + save_filepath
        save_csv_name = savefp +  str(self.file_name) + 'cycleRate'
        if not os.path.exists(savefp):
            os.makedirs(savefp)
        if not os.path.exists(save_csv_name + '.csv'):
            with open(save_csv_name + '.csv', 'w') as f:
                writer = csv.writer(f, delimiter = '\t')
                writer.writerows(cycleRate)


class Analysis_LPTS:
    LE_LEBC = 3.25 #m
    HE_HEBC = 6.71
    # the first for S and the second for A ring
    LEBC_ring = [7.24,9.59] #m
    HEBC_ring = [9.93,7.80]
    ring_circ = [8.68,8.71] # m
    rad = ring_circ[0]/(2*np.pi) #m
    cen_RAES = 0
    light_speed = 299792458
    
    def __init__(self, source_path, saving_path, multiple_file_array,molecule,ion_ke):
        # self.q = queue.Queue()    
        self.DES_path = source_path
        self.save_path = saving_path
        self.laser_on = False
        self.multiple_file_array = multiple_file_array
        molecule_amu = formula('C[12]60').mass
        ion_mass = molecule_amu * const.atomic_mass
        self.ion_velocity = np.sqrt(2 * ion_ke * const.value('electron volt') / ion_mass)
        if not os.path.exists(self.save_path):
            os.makedirs(self.save_path)
            print("Directory " , self.save_path,  " Created ")
        else:    
            print("Directory " , self.save_path ,  " already exists")


    def ini_params(self,exp, _file,start,stop,hdf5_groupname, n1,n2,bins,dump):
        self.file_name = _file
        self.experiment = exp
        self.bins = bins
        self.start = start
        self.stop = stop
        self.cycle_max = n2
        self.cycle_min = n1
        self.beam_dump = dump
        self.savePath = self.save_path + self.experiment + self.file_name + '.hdf'
        self.tdms_path = self.DES_path + self.experiment + self.file_name + '.tdms'
        # self.q.put(self.tdms_path)
        self.Signal_vs_power()

      
    def Signal_vs_power(self):
        self.atime = time()
        with TdmsFile.open(self.tdms_path) as tdms_file:
            # print (tdms_file.properties)
            group= tdms_file.groups()
        #     # if gr.properties['Description'] == self.channel_names[self.No_Chan[0]]: # the name of the channel might change
            times = group[5]['Time (s)']
            n_cycle = group[5]['Cycle No.']
            n_c = len(times)
            m_c = round(max(times),6)
            number_cycles = len(n_cycle)
            print ('from file: ', self.file_name[-7:], ', length of channel: ', n_c, 'cycle time (rounded):', m_c)
            self.b_norm = self.bin_data(times,self.beam_dump,self.beam_dump + 7e-2)
            back = 0
            for i in range(len(self.start)):
                signal = self.bin_data(times,self.start[i],self.stop[i])
                back += self.bin_data(times,self.stop[i] ,self.stop[i] + 24e-2)
                self.integrated_shots[self.file_name[-7:]][0] += (signal - self.b_norm/70)               
                self.integrated_shots[self.file_name[-7:]][1] += np.sqrt((signal + self.b_norm/70))
            self.integrated_shots[self.file_name[-7:]][2] = back/number_cycles
            self.integrated_shots[self.file_name[-7:]][3] = np.sqrt(back)/number_cycles

        print (' Handling of : ', self.file_name[-7:], ' took : ' + str(round(time() - self.atime,4)) + ' s ')

        # print ('All took : ' + str(round(time() - self.atime,4)) + ' s ')

    def bin_data(self,data,start,stop):
        his, bin_edges = np.histogram(data, bins = self.bins, range=(start, stop))
        return his

        
    def doppler_shift(self,lam0_,theta,co):
        ''' not  relativistic'''
        c = self.light_speed
        lam0 = lam0_*1e-9
        nu_0 = c/lam0
        h_eV = 4.135667696e-15
        nu_p_co = nu_0*(1 - self.ion_velocity/c*np.cos(theta*np.pi/180))
        nu_p = nu_0*(1 + self.ion_velocity/c*np.cos(theta*np.pi/180))            
        # print ('nu: ' ,round((c/nu_p)*1e9,2), 'nu0: ', round((c/nu_0)*1e9,2))
        return nu_p*h_eV, nu_p_co*h_eV
            
    '''
    Saving function1
    '''
        
    def save_data(self, z_sub, backgnd,c_rate,wavelengths):
        savefp = self.save_path + '/' + self.file_name
        if not os.path.exists(savefp + '/pars/'):
            os.makedirs(savefp + '/pars/')
            
        if self.laser_on:
            save_csv_name = savefp  + '/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' +str(self.l)
            save_csv_pars = savefp + '/pars/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' + str(len(wavelengths)) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_' + str(len(wavelengths)) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))

            for i in range(len(wavelengths)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}.csv'.format(wavelengths[i]), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])

        else:
            save_csv_name = savefp  + '/' + self.method + '_' + str(self.file_name)
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
        
            for i in range(len(self.channels)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}_{}_countRate.csv'.format(self.channels[i],self.l), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])
        print ('saving' + str(wavelengths) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')

        with open(save_csv_name + '_{}_{}_pars.csv'.format(self.channels[i],self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)
            
            
    def save_data_OPO_pow(self, z_sub, backgnd):
        savefp = self.save_path + '/Qsw'
        save_csv_name = savefp  + '/' + self.method + '_' + str(self.file_name)

        if not os.path.exists(savefp):
            os.makedirs(savefp)
        
        for i in range(len(self.channels)):
            # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
            with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                writer = csv.writer(arch, delimiter = '\t')
                writer.writerows(z_sub[i])
            print ('file created in:', savefp, ' under the name ', self.method + '_' + str(self.file_name) + '_{}_{}.csv'.format(self.channels[i],self.l))
            
        with open(save_csv_name + '_{}_{}_pars.csv'.format(self.channels[i],self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)
            
        print ('saving' + str(self.file_name) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')

            
            
    def save_data_LPTS(self,z,pows):
        savefp = self.save_path + '/LPTS/'
        if not os.path.exists(savefp):
            os.makedirs(savefp)
        print ('The shape of your data: ', np.shape(z))
        with open(savefp + 'LPTS{}.csv'.format(self.file_name), 'w') as arch:
            writer = csv.writer(arch, delimiter = '\t')
            writer.writerows(z)
        with open(savefp + 'LPTS{}_pows.csv'.format(self.file_name), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(pows)
            
    def save_pickle_data(self, z_sub, backgnd,c_rate,wavelengths):
        savefp = self.save_path + '/' + self.file_name
        if not os.path.exists(savefp):
            os.makedirs(savefp)
            
        if not os.path.exists(savefp + '/pars/'):
            os.makedirs(savefp + '/pars/')
        
        # if not os.path.exists(savefp + 'countRates/'):
            # print('making folder: ' + savefp + 'countRates/')
            # os.makedirs(savefp + 'countRates/')
            
            
        if self.laser_on:
            save_csv_name = savefp  + '/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' +str(self.l)
            save_csv_pars = savefp + '/pars/' + str(self.file_name) + '_' + str(self.channels[self.counter]) + '_' + str(len(wavelengths)) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_' + str(len(wavelengths)) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
    
            for i in range(len(wavelengths)):
                for j in range(len(self.channels)):
                    # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                    with open(save_csv_name + '_{}_{}.pickle'.format(self.channels[j],wavelengths[i]), 'wb') as arch:
                        try:
                            writer = pickle.dump(z_sub[i],arch, protocol = pickle.HIGHEST_PROTOCOL)
                        except Exception as ex:
                            print ('Error during pickling: ', ex)
    
                    if self.iscountsOn:
                        with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                            writer = csv.writer(arch, delimiter = '\t')
                            writer.writerows(c_rate[i])
    
        else:
            save_csv_name = savefp  + self.method + str(self.file_name) + str(self.l)
            save_csv_pars = savefp + 'pars/' + self.method + str(self.file_name) + '_pars'
            save_csv_cRate = savefp + 'countRates/' + str(self.file_name) + '_countRates'
            print ('file created in:', savefp, ' under the name ', str(self.file_name))
        
            for i in range(len(self.channels)):
                # if not os.path.exists(save_csv_name + '_{}.csv'.format(self.channels[i])):
                with open(save_csv_name + '_{}_{}.csv'.format(self.channels[i],self.l), 'w') as arch:
                    writer = csv.writer(arch, delimiter = '\t')
                    writer.writerows(z_sub[i])
                if self.iscountsOn:
                    with open(save_csv_cRate + '_{}.csv'.format(self.channels[i]), 'w') as arch:
                        writer = csv.writer(arch, delimiter = '\t')
                        writer.writerows(c_rate[i])
        print ('saving' + str(wavelengths) + ' took : ' + str(round(time() - self.atime,4)) + ' s ')
    
    
        with open(save_csv_pars + '{}.csv'.format(self.l), 'w') as b_arch:
            writer = csv.writer(b_arch, delimiter = '\t')
            writer.writerows(backgnd)

